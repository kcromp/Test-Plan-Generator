
import test_plan_gui

import openpyxl 

formText = [
'Special characters',
'Blank text',
'Enter word NULL',
'Enter characters past the max amount allowed'
]

formTextLen = len(formText)
wb = openpyxl.load_workbook(test_plan_gui.path)
sheet = wb.active

for i in range(1,sheet.max_row+1):
	if (sheet.cell(row=i, column=1).value) == "form text ":
		sheet.insert_rows(idx=i, amount=formTextLen-1)

		for row, text in enumerate(formText, start=i):
			sheet.cell(column=1, row=row, value=text)
			wb.save(test_plan_gui.path)