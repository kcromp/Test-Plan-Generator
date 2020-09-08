import test_plan_gui

import openpyxl 

formSubmit = [
'Submit confirmation',
'Check front end display of newly changed information',
'Verify update to SQL',
'Refresh page',
'Hit Submit again'
]

formSubmitLen = len(formSubmit)
wb = openpyxl.load_workbook(test_plan_gui.path)
sheet = wb.active

for i in range(1,sheet.max_row+1):
	if (sheet.cell(row=i, column=1).value) == "form submit ":
		sheet.insert_rows(idx=i, amount=formSubmitLen-1)

		for row, text in enumerate(formSubmit, start=i):
			sheet.cell(column=1, row=row, value=text)
			wb.save(test_plan_gui.path)

import form_text